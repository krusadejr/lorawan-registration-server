import os
from flask import Flask, render_template, request, redirect, url_for, flash, session, Response, stream_with_context, send_file
import pandas as pd
from werkzeug.utils import secure_filename
import uuid
import json
import logging
from datetime import datetime
from file_parser import parse_file, get_column_info
import time
import io
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this-in-production'
app.config['TEMPLATES_AUTO_RELOAD'] = True  # Disable Jinja2 template caching

# Configuration
UPLOAD_FOLDER = 'uploads'
LOG_FOLDER = 'logs'
CONFIG_HISTORY_FILE = 'config_history.json'
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'xlsm', 'txt', 'json', 'csv'}
MAX_FILE_SIZE = 16 * 1024 * 1024  # 16MB max file size
MAX_HISTORY_ITEMS = 5  # Maximum number of items to keep in history

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

# Create necessary directories
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(LOG_FOLDER, exist_ok=True)

# Setup logging
log_filename = os.path.join(LOG_FOLDER, f'app_{datetime.now().strftime("%Y%m%d")}.log')
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename),
        logging.StreamHandler()  # Also print to console
    ]
)
logger = logging.getLogger(__name__)
logger.info("="*80)
logger.info("Application started")
logger.info("="*80)

# Global variables to store server configuration
SERVER_URL = None               # ChirpStack server URL
API_CODE = None                 # API key for authentication
TENANT_ID = None                # Tenant ID


def _is_valid_uuid(uuid_string):
    """
    Validate if a string is a valid UUID format.
    UUID should be 36 characters: XXXXXXXX-XXXX-XXXX-XXXX-XXXXXXXXXXXX
    """
    try:
        if not isinstance(uuid_string, str):
            return False
        uuid_string = uuid_string.strip()
        # Try to parse it as UUID
        uuid.UUID(uuid_string)
        return True
    except (ValueError, AttributeError):
        return False


def cleanup_upload_cache(keep_count=20):
    """
    Clean up old upload files, keeping only the last N files.
    This prevents unlimited disk space usage from accumulated uploads.
    """
    try:
        if not os.path.exists(UPLOAD_FOLDER):
            return 0, 0
        
        # Get all files in upload folder
        all_files = []
        for filename in os.listdir(UPLOAD_FOLDER):
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            if os.path.isfile(filepath):
                # Get modification time
                mtime = os.path.getmtime(filepath)
                all_files.append((filepath, mtime, filename))
        
        # If we have more than keep_count files, delete the oldest ones
        if len(all_files) > keep_count:
            # Sort by modification time (newest first)
            all_files.sort(key=lambda x: x[1], reverse=True)
            
            # Files to delete are those beyond keep_count
            files_to_delete = all_files[keep_count:]
            deleted_count = 0
            total_freed = 0
            
            for filepath, mtime, filename in files_to_delete:
                try:
                    file_size = os.path.getsize(filepath)
                    os.remove(filepath)
                    total_freed += file_size
                    deleted_count += 1
                    logger.info(f"Deleted old cache file: {filename} ({file_size} bytes)")
                except Exception as e:
                    logger.error(f"Error deleting file {filename}: {e}")
            
            if deleted_count > 0:
                logger.info(f"Cache cleanup: Deleted {deleted_count} files, freed {total_freed / 1024 / 1024:.2f} MB")
            
            return deleted_count, total_freed
        
        return 0, 0
    
    except Exception as e:
        logger.error(f"Error during cache cleanup: {e}")
        return 0, 0


def get_upload_cache_status():
    """Get current upload cache statistics."""
    try:
        if not os.path.exists(UPLOAD_FOLDER):
            return {'file_count': 0, 'total_size': 0, 'size_mb': 0}
        
        file_count = 0
        total_size = 0
        
        for filename in os.listdir(UPLOAD_FOLDER):
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            if os.path.isfile(filepath):
                file_count += 1
                total_size += os.path.getsize(filepath)
        
        return {
            'file_count': file_count,
            'total_size': total_size,
            'size_mb': round(total_size / 1024 / 1024, 2)
        }
    except Exception as e:
        logger.error(f"Error getting cache status: {e}")
        return {'file_count': 0, 'total_size': 0, 'size_mb': 0}


# Run cache cleanup on startup (keep last 20 files)
logger.info("Running upload cache cleanup...")
deleted, freed = cleanup_upload_cache(keep_count=20)
if deleted > 0:
    logger.info(f"Cleanup on startup: Deleted {deleted} old files, freed {freed / 1024 / 1024:.2f} MB")
else:
    logger.info("Upload cache OK - within retention limit")


def load_config_history():
    """Load configuration history from JSON file."""
    if os.path.exists(CONFIG_HISTORY_FILE):
        try:
            with open(CONFIG_HISTORY_FILE, 'r') as f:
                return json.load(f)
        except Exception as e:
            logger.error(f"Error loading config history: {e}")
            return {'server_urls': [], 'api_keys': [], 'tenant_ids': []}
    return {'server_urls': [], 'api_keys': [], 'tenant_ids': []}


def save_config_history(history):
    """Save configuration history to JSON file."""
    try:
        with open(CONFIG_HISTORY_FILE, 'w') as f:
            json.dump(history, f, indent=2)
        logger.info("Config history saved successfully")
    except Exception as e:
        logger.error(f"Error saving config history: {e}")


def add_to_history(history, key, value):
    """Add a value to history list, maintaining max size and avoiding duplicates."""
    if not value or not value.strip():
        return history
    
    value = value.strip()
    
    # Initialize list if not exists
    if key not in history:
        history[key] = []
    
    # Remove if already exists (to move it to front)
    if value in history[key]:
        history[key].remove(value)
    
    # Add to front
    history[key].insert(0, value)
    
    # Limit to MAX_HISTORY_ITEMS
    history[key] = history[key][:MAX_HISTORY_ITEMS]
    
    return history


def allowed_file(filename):
    """Check if the uploaded file has an allowed extension."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

 
@app.route('/')
def index():
    """Home page with file upload form."""
    global SERVER_URL, API_CODE, TENANT_ID
    history = load_config_history()
    cache_status = get_upload_cache_status()
    return render_template('index.html', 
                         server_url=SERVER_URL,
                         api_code=API_CODE,
                         tenant_id=TENANT_ID,
                         device_profile_id=None,
                         history=history,
                         cache_status=cache_status)


@app.route('/server-config')
def server_config():
    """Server configuration page."""
    global SERVER_URL, API_CODE, TENANT_ID
    history = load_config_history()
    return render_template('server_config.html', 
                         server_url=SERVER_URL, 
                         api_code=API_CODE,
                         tenant_id=TENANT_ID,
                         device_profile_id=None,
                         history=history)


@app.route('/save-server-config', methods=['POST'])
def save_server_config():
    """Save server configuration to global variables."""
    global SERVER_URL, API_CODE, TENANT_ID
    
    server_url = request.form.get('server_url', '').strip()
    api_code = request.form.get('api_code', '').strip()
    tenant_id = request.form.get('tenant_id', '').strip()
    
    logger.info(f"Saving server config: server_url='{server_url}', api_code_length={len(api_code)}, tenant_id='{tenant_id}'")
    
    if not any([server_url, api_code, tenant_id]):
        flash('Bitte geben Sie mindestens einen Wert ein', 'danger')
        return redirect(url_for('server_config'))
    
    saved_vars = []
    
    if server_url:
        SERVER_URL = server_url
        saved_vars.append('SERVER_URL')
        logger.info(f"Set SERVER_URL to: {SERVER_URL}")
    
    if api_code:
        API_CODE = api_code
        saved_vars.append('API_CODE')
        logger.info(f"Set API_CODE (length: {len(API_CODE)})")
    
    if tenant_id:
        TENANT_ID = tenant_id
        saved_vars.append('TENANT_ID')
        logger.info(f"Set TENANT_ID to: {TENANT_ID}")
    
    logger.info(f"Global variables after save: SERVER_URL={SERVER_URL}, API_CODE length={len(API_CODE) if API_CODE else 0}, TENANT_ID={TENANT_ID}")
    
    # Save to history
    history = load_config_history()
    if server_url:
        history = add_to_history(history, 'server_urls', server_url)
    if api_code:
        history = add_to_history(history, 'api_keys', api_code)
    if tenant_id:
        history = add_to_history(history, 'tenant_ids', tenant_id)
    save_config_history(history)
    
    flash(f'Server-Konfiguration erfolgreich gespeichert: {", ".join(saved_vars)}', 'success')
    
    # Redirect to index with a prompt to upload file
    flash('✓ Server konfiguriert! Sie können jetzt Geräte-Dateien hochladen.', 'info')
    return redirect(url_for('index'))


@app.route('/clear-history', methods=['POST'])
def clear_history():
    """Clear configuration history."""
    try:
        if os.path.exists(CONFIG_HISTORY_FILE):
            os.remove(CONFIG_HISTORY_FILE)
        logger.info("Config history cleared")
        flash('Konfigurationsverlauf erfolgreich gelöscht', 'success')
    except Exception as e:
        logger.error(f"Error clearing history: {e}")
        flash(f'Fehler beim Löschen des Verlaufs: {str(e)}', 'danger')
    return redirect(url_for('index'))


@app.route('/last-sessions')
def last_sessions():
    """Display last accessed server configurations."""
    global SERVER_URL, API_CODE, TENANT_ID
    
    logger.info("="*80)
    logger.info("LAST SESSIONS REQUEST")
    logger.info("="*80)
    
    history = load_config_history()
    logger.info(f"Loaded config history: {history}")
    
    # Create session list with combined configurations
    sessions = []
    
    # Get the maximum number of sessions from the history
    max_sessions = max(
        len(history.get('server_urls', [])),
        len(history.get('api_keys', [])),
        len(history.get('tenant_ids', []))
    )
    
    for i in range(min(max_sessions, MAX_HISTORY_ITEMS)):
        session_data = {
            'id': i,
            'server_url': history.get('server_urls', [])[i] if i < len(history.get('server_urls', [])) else '',
            'api_key': history.get('api_keys', [])[i] if i < len(history.get('api_keys', [])) else '',
            'tenant_id': history.get('tenant_ids', [])[i] if i < len(history.get('tenant_ids', [])) else '',
        }
        sessions.append(session_data)
    
    logger.info(f"Prepared {len(sessions)} sessions for display")
    
    return render_template('last_sessions.html',
                         sessions=sessions,
                         current_server_url=SERVER_URL,
                         current_api_key=API_CODE,
                         current_tenant_id=TENANT_ID)


@app.route('/load-session/<int:session_id>', methods=['POST'])
def load_session(session_id):
    """Load a saved server configuration session."""
    global SERVER_URL, API_CODE, TENANT_ID
    
    logger.info("="*80)
    logger.info(f"LOAD SESSION REQUEST: session_id={session_id}")
    logger.info("="*80)
    
    history = load_config_history()
    
    # Validate session_id
    if session_id >= len(history.get('server_urls', [])):
        logger.error(f"Invalid session_id: {session_id}")
        flash('Ungültige Session-ID', 'danger')
        return redirect(url_for('last_sessions'))
    
    # Load session data
    server_url = history.get('server_urls', [])[session_id] if session_id < len(history.get('server_urls', [])) else ''
    api_key = history.get('api_keys', [])[session_id] if session_id < len(history.get('api_keys', [])) else ''
    tenant_id = history.get('tenant_ids', [])[session_id] if session_id < len(history.get('tenant_ids', [])) else ''
    
    # Set global variables
    if server_url:
        SERVER_URL = server_url
        logger.info(f"Loaded SERVER_URL: {SERVER_URL}")
    if api_key:
        API_CODE = api_key
        logger.info(f"Loaded API_CODE (length: {len(API_CODE)})")
    if tenant_id:
        TENANT_ID = tenant_id
        logger.info(f"Loaded TENANT_ID: {TENANT_ID}")
    
    flash(f'Session #{session_id + 1} geladen! Server-Konfiguration wiederhergestellt.', 'success')
    return redirect(url_for('index'))


@app.route('/test-connection')
def test_connection():
    """Test connection to ChirpStack server."""
    global SERVER_URL, API_CODE
    
    logger.info("="*80)
    logger.info("TEST CONNECTION REQUEST")
    logger.info("="*80)
    
    # Check if server is configured
    if not SERVER_URL or not API_CODE:
        logger.warning("Server not configured")
        return render_template('test_connection.html', 
                             configured=False,
                             server_url=SERVER_URL,
                             api_code=API_CODE)
    
    logger.info(f"Testing connection to: {SERVER_URL}")
    
    # Try to connect
    connection_result = {
        'success': False,
        'message': '',
        'details': []
    }
    
    try:
        from grpc_client import ChirpStackClient
        
        # Create client
        client = ChirpStackClient(SERVER_URL, API_CODE)
        connection_result['details'].append('✓ Client erstellt')
        
        # Test connection
        connected, conn_msg = client.connect()
        connection_result['details'].append(f'Verbindungsversuch: {conn_msg}')
        
        if connected:
            connection_result['success'] = True
            connection_result['message'] = 'Erfolgreich mit ChirpStack verbunden!'
            logger.info("Connection test successful")
        else:
            connection_result['message'] = f'Verbindung fehlgeschlagen: {conn_msg}'
            logger.error(f"Connection test failed: {conn_msg}")
        
        # Close connection
        client.close()
        
    except ImportError as e:
        connection_result['message'] = f'gRPC Client konnte nicht geladen werden: {str(e)}'
        connection_result['details'].append(f'✗ Import-Fehler: {str(e)}')
        logger.error(f"Import error: {e}")
    except Exception as e:
        connection_result['message'] = f'Fehler beim Verbindungstest: {str(e)}'
        connection_result['details'].append(f'✗ Fehler: {str(e)}')
        logger.error(f"Connection test error: {e}", exc_info=True)
    
    return render_template('test_connection.html',
                         configured=True,
                         server_url=SERVER_URL,
                         api_code=API_CODE,
                         result=connection_result)


@app.route('/lorawan-version-detector')
def lorawan_version_detector():
    """
    Diagnostic page to detect LoRaWAN versions of device profiles
    Shows which device profiles are 1.0.x, 1.1.x, OTAA, ABP, etc.
    """
    import logging
    logger = logging.getLogger(__name__)
    
    version_info = {
        'configured': False,
        'server_url': SERVER_URL,
        'tenant_id': TENANT_ID,
        'profiles': [],
        'errors': [],
        'summary': {}
    }
    
    try:
        if not SERVER_URL or not API_CODE:
            version_info['errors'].append("ChirpStack server nicht konfiguriert")
            return render_template('lorawan_version_detector.html', **version_info)
        
        logger.info("[Diagnostics] Fetching device profiles and LoRaWAN versions...")
        
        # Connect and fetch profiles
        from grpc_client import ChirpStackClient
        client = ChirpStackClient(SERVER_URL, API_CODE)
        client.connect()
        
        success, result = client.get_device_profiles_via_rest(TENANT_ID)
        
        if not success:
            version_info['errors'].append(f"Fehler beim Abrufen von Geräteprofilen: {result}")
            logger.error(f"[Diagnostics] Failed to fetch profiles: {result}")
        else:
            version_info['configured'] = True
            version_info['profiles'] = result
            
            # Generate summary
            v100_count = sum(1 for p in result if p['mac_version']['minor'] == 0)
            v110_count = sum(1 for p in result if p['mac_version']['minor'] == 1)
            otaa_count = sum(1 for p in result if p['supports_otaa'])
            abp_count = sum(1 for p in result if p['supports_abp'])
            
            version_info['summary'] = {
                'total_profiles': len(result),
                'lorawan_1_0_x': v100_count,
                'lorawan_1_1_x': v110_count,
                'supports_otaa': otaa_count,
                'supports_abp': abp_count
            }
            
            logger.info(f"[Diagnostics] Found {len(result)} device profiles: {v100_count} x 1.0.x, {v110_count} x 1.1.x")
        
        client.close()
        
    except ImportError as e:
        version_info['errors'].append(f"Import-Fehler: {str(e)}")
        logger.error(f"[Diagnostics] Import error: {e}")
    except Exception as e:
        version_info['errors'].append(f"Fehler: {str(e)}")
        logger.error(f"[Diagnostics] Error: {e}", exc_info=True)
    
    return render_template('lorawan_version_detector.html', **version_info)


@app.route('/help')
def help_page():
    """Help page with instructions for getting ChirpStack IDs."""
    return render_template('help.html')


@app.route('/download-template')
def download_template():
    """Download Excel template with correct column headers and example data."""
    logger.info("Template download requested")
    
    # Create template DataFrame with example data including tags
    template_data = {
        'dev_eui': ['0000000000000001', '0000000000000002', '0000000000000003', '0000000000000004', '0000000000000005'],
        'name': ['Sensor_Floor_1', 'Sensor_Floor_2', 'Humidity_Room_A', 'Motion_Corridor', 'Light_Sensor_Main'],
        'application_id': ['app-uuid-12345678', 'app-uuid-12345678', 'app-uuid-12345678', 'app-uuid-12345678', 'app-uuid-12345678'],
        'device_profile_id': [
            'profile-uuid-87654321',
            'profile-uuid-87654321',
            'profile-uuid-87654321',
            'profile-uuid-87654321',
            'profile-uuid-87654321'
        ],
        'nwk_key': [
            '00112233445566778899AABBCCDDEEFF',
            '11223344556677889900AABBCCDDEEFF',
            '22334455667788990011AABBCCDDEEFF',
            '33445566778899001122AABBCCDDEEFF',
            '44556677889900112233AABBCCDDEEFF'
        ],
        'app_key': [
            '00112233445566778899AABBCCDDEEFF',
            '11223344556677889900AABBCCDDEEFF',
            '22334455667788990011AABBCCDDEEFF',
            '33445566778899001122AABBCCDDEEFF',
            '44556677889900112233AABBCCDDEEFF'
        ],
        'description': [
            'Temperature sensor in floor 1',
            'Temperature sensor in floor 2',
            'Humidity sensor in room A',
            'Motion detector in main corridor',
            'Light level sensor in main area'
        ],
        'tags': [
            'location:floor1|type:temp|status:active',
            'location:floor2|type:temp|status:active',
            'location:roomA|type:humidity|status:active',
            'location:corridor|type:motion|status:active',
            'location:main|type:light|status:active'
        ]
    }
    
    df = pd.DataFrame(template_data)
    
    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Devices', index=False)
        
        # Auto-adjust column widths
        worksheet = writer.sheets['Devices']
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='device_registration_template.xlsx'
    )


@app.route('/export-server-config')
def export_server_config():
    """Export server configuration details as a text file."""
    logger.info("Server configuration export requested")
    
    if not SERVER_URL:
        flash('Server-Konfiguration nicht vollständig. Bitte zuerst konfigurieren.', 'error')
        return redirect(url_for('server_config'))
    
    # Create configuration content - simple format with full API key
    config_content = f"""LoRaWAN Registration Server - Configuration Details
{'=' * 60}
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

SERVER INFORMATION
{'-' * 60}
Server URL:     {SERVER_URL}
Tenant ID:      {TENANT_ID if TENANT_ID else 'Not configured'}
API Key:        {API_CODE if API_CODE else 'Not configured'}
"""
    
    # Create file in memory
    output = io.BytesIO()
    output.write(config_content.encode('utf-8'))
    output.seek(0)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"ChirpStack_Configuration_{timestamp}.txt"
    
    logger.info(f"Exporting server configuration: {filename}")
    
    return send_file(
        output,
        mimetype='text/plain',
        as_attachment=True,
        download_name=filename
    )

@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle file upload and parse data."""
    logger.info("="*80)
    logger.info("UPLOAD FILE REQUEST")
    logger.info("="*80)
    
    # Check if file is present in request
    if 'file' not in request.files:
        logger.warning("No file in request")
        flash('Keine Datei in der Anfrage', 'danger')
        return redirect(url_for('index'))
    
    file = request.files['file']
    logger.info(f"File received: {file.filename}")
    
    # Check if user selected a file
    if file.filename == '':
        logger.warning("Empty filename")
        flash('Keine Datei ausgewählt', 'danger')
        return redirect(url_for('index'))
    
    # Check if file is allowed
    if file and allowed_file(file.filename):
        # Generate unique filename to avoid conflicts
        unique_id = str(uuid.uuid4())
        original_filename = secure_filename(file.filename)
        file_extension = original_filename.rsplit('.', 1)[1].lower()
        unique_filename = f"{unique_id}.{file_extension}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        
        logger.info(f"Saving file to: {filepath}")
        file.save(filepath)
        
        try:
            # Parse file using our parser
            logger.info(f"Parsing file with extension: {file_extension}")
            parse_result = parse_file(filepath, file_extension)
            
            logger.info(f"Parse result success: {parse_result['success']}")
            logger.info(f"Parse result message: {parse_result['message']}")
            
            if not parse_result['success']:
                flash(parse_result['message'], 'danger')
                if os.path.exists(filepath):
                    os.remove(filepath)
                return redirect(url_for('index'))
            
            logger.info(f"Number of sheets: {len(parse_result['sheets'])}")
            logger.info(f"Sheet names: {parse_result['sheets']}")
            
            # Store in session - only metadata, not the actual data
            session['filepath'] = filepath
            session['original_filename'] = original_filename
            session['file_type'] = parse_result['file_type']
            session['sheet_names'] = list(parse_result['sheets'])  # Ensure it's a list
            
            # Save parsed data to a temporary JSON file instead of session
            parsed_data_file = os.path.join(app.config['UPLOAD_FOLDER'], f"{unique_id}_parsed.json")
            session_data = {}
            for sheet_name, df in parse_result['data'].items():
                session_data[sheet_name] = df.to_dict(orient='records')
                logger.info(f"Sheet '{sheet_name}': {len(df)} rows, {len(df.columns)} columns")
            
            logger.info(f"Writing parsed data to: {parsed_data_file}")
            with open(parsed_data_file, 'w') as f:
                json.dump(session_data, f)
            
            session['parsed_data_file'] = parsed_data_file
            
            # Log session state
            logger.info(f"Session stored - sheet_names: {session.get('sheet_names')}")
            logger.info(f"Session stored - parsed_data_file: {parsed_data_file}")
            logger.info(f"Session stored - filepath: {filepath}")
            logger.info(f"File exists check: {os.path.exists(parsed_data_file)}")
            
            # Check if delimiter input is needed
            if parse_result.get('needs_delimiter', False):
                session['needs_delimiter'] = True
                session['delimiter_info'] = parse_result.get('delimiter_info', {})
                flash('Bitte geben Sie das Trennzeichen für die Datei an', 'info')
                logger.info("Redirecting to delimiter_input")
                return redirect(url_for('delimiter_input'))
            
            flash(parse_result['message'], 'success')
            
            # Redirect to sheet selection page
            logger.info("Redirecting to select_sheet")
            return redirect(url_for('select_sheet'))
        
        except Exception as e:
            logger.error(f"Error processing file: {str(e)}", exc_info=True)
            flash(f'Fehler beim Verarbeiten der Datei: {str(e)}', 'danger')
            # Clean up file if error occurs
            if os.path.exists(filepath):
                os.remove(filepath)
            return redirect(url_for('index'))
    
    else:
        logger.warning(f"Invalid file type: {file.filename}")
        flash('Ungültiger Dateityp. Bitte laden Sie eine gültige Datei hoch (.xlsx, .xls, .xlsm, .csv, .txt, .json)', 'danger')
        return redirect(url_for('index'))


@app.route('/delimiter-input')
def delimiter_input():
    """Page for manual delimiter input when auto-detection fails."""
    logger.info("="*80)
    logger.info("DELIMITER INPUT REQUEST")
    logger.info("="*80)
    
    if not session.get('needs_delimiter', False):
        logger.warning("Delimiter input not needed, redirecting to select_sheet")
        return redirect(url_for('select_sheet'))
    
    original_filename = session.get('original_filename', '')
    delimiter_info = session.get('delimiter_info', {})
    
    logger.info(f"Delimiter input needed for: {original_filename}")
    logger.info(f"Delimiter info: {delimiter_info}")
    
    return render_template('delimiter_input.html',
                         original_filename=original_filename,
                         delimiter_info=delimiter_info)


@app.route('/process-delimiter', methods=['POST'])
def process_delimiter():
    """Process the user-provided delimiter and re-parse the file."""
    logger.info("="*80)
    logger.info("PROCESS DELIMITER REQUEST")
    logger.info("="*80)
    
    delimiter = request.form.get('delimiter', '').strip()
    custom_delimiter = request.form.get('custom_delimiter', '').strip()
    
    # Use custom delimiter if provided, otherwise use the selected one
    if custom_delimiter:
        delimiter = custom_delimiter
        logger.info(f"Using custom delimiter: repr={repr(delimiter)}")
    else:
        logger.info(f"Using predefined delimiter: {delimiter}")
    
    if not delimiter:
        flash('Bitte wählen Sie ein Trennzeichen aus oder geben Sie ein eigenes ein', 'danger')
        return redirect(url_for('delimiter_input'))
    
    # Convert common delimiter names to actual characters
    delimiter_map = {
        'comma': ',',
        'semicolon': ';',
        'tab': '\t',
        'pipe': '|',
        'space': ' '
    }
    
    actual_delimiter = delimiter_map.get(delimiter, delimiter)
    logger.info(f"Actual delimiter to use: repr={repr(actual_delimiter)}")
    
    # Get file info from session
    filepath = session.get('filepath')
    file_extension = filepath.rsplit('.', 1)[1].lower() if filepath else None
    
    if not filepath or not os.path.exists(filepath):
        flash('Datei nicht gefunden. Bitte laden Sie die Datei erneut hoch.', 'danger')
        return redirect(url_for('index'))
    
    try:
        # Re-parse with the specified delimiter
        from file_parser import parse_csv_txt_with_delimiter
        parse_result = parse_csv_txt_with_delimiter(filepath, actual_delimiter)
        
        if not parse_result['success']:
            flash(f"Fehler beim Parsen mit dem angegebenen Trennzeichen: {parse_result['message']}", 'danger')
            return redirect(url_for('delimiter_input'))
        
        # Update session data
        unique_id = filepath.rsplit('.', 1)[0].rsplit(os.sep, 1)[1]
        parsed_data_file = os.path.join(app.config['UPLOAD_FOLDER'], f"{unique_id}_parsed.json")
        
        session_data = {}
        for sheet_name, df in parse_result['data'].items():
            session_data[sheet_name] = df.to_dict(orient='records')
            logger.info(f"Sheet '{sheet_name}': {len(df)} rows, {len(df.columns)} columns")
        
        with open(parsed_data_file, 'w') as f:
            json.dump(session_data, f)
        
        session['parsed_data_file'] = parsed_data_file
        session['sheet_names'] = list(parse_result['sheets'])
        session['needs_delimiter'] = False
        session.pop('delimiter_info', None)
        
        flash(parse_result['message'], 'success')
        logger.info("Delimiter processing successful, redirecting to select_sheet")
        return redirect(url_for('select_sheet'))
        
    except Exception as e:
        logger.error(f"Error processing delimiter: {str(e)}", exc_info=True)
        flash(f'Fehler beim Verarbeiten: {str(e)}', 'danger')
        return redirect(url_for('delimiter_input'))


@app.route('/select-sheet')
def select_sheet():
    """Sheet selection page."""
    logger.info("="*80)
    logger.info("SELECT SHEET REQUEST")
    logger.info("="*80)
    
    sheet_names = session.get('sheet_names', [])
    original_filename = session.get('original_filename', '')
    file_type = session.get('file_type', '')
    parsed_data_file = session.get('parsed_data_file', '')
    
    logger.info(f"Session sheet_names: {sheet_names}")
    logger.info(f"Session original_filename: {original_filename}")
    logger.info(f"Session file_type: {file_type}")
    logger.info(f"Session parsed_data_file: {parsed_data_file}")
    logger.info(f"All session keys: {list(session.keys())}")
    
    if parsed_data_file:
        file_exists = os.path.exists(parsed_data_file)
        logger.info(f"Parsed data file exists: {file_exists}")
    else:
        logger.warning("No parsed_data_file in session")
    
    if not sheet_names or not parsed_data_file or not os.path.exists(parsed_data_file):
        logger.error("Missing data - redirecting to index")
        logger.error(f"  sheet_names empty: {not sheet_names}")
        logger.error(f"  parsed_data_file empty: {not parsed_data_file}")
        logger.error(f"  file exists: {os.path.exists(parsed_data_file) if parsed_data_file else False}")
        flash('Keine Daten gefunden. Bitte laden Sie die Datei erneut hoch.', 'danger')
        return redirect(url_for('index'))
    
    # Read parsed data from file
    logger.info(f"Reading parsed data from: {parsed_data_file}")
    with open(parsed_data_file, 'r') as f:
        parsed_data = json.load(f)
    
    logger.info(f"Parsed data contains sheets: {list(parsed_data.keys())}")
    
    # Get preview data for each sheet
    sheet_previews = {}
    
    for sheet_name in sheet_names:
        if sheet_name in parsed_data:
            df = pd.DataFrame(parsed_data[sheet_name])
            logger.info(f"Creating preview for sheet '{sheet_name}': {len(df)} rows, {len(df.columns)} cols")
            sheet_previews[sheet_name] = {
                'rows': len(df),
                'columns': len(df.columns),
                'column_names': list(df.columns),
                'preview_html': df.head(5).to_html(
                    classes='table is-bordered is-striped is-hoverable is-fullwidth',
                    index=False,
                    na_rep='N/A'
                )
            }
        else:
            logger.warning(f"Sheet '{sheet_name}' not found in parsed data")
    
    return render_template('select_sheet.html',
                         filename=original_filename,
                         file_type=file_type,
                         sheet_names=sheet_names,
                         sheet_previews=sheet_previews)


@app.route('/column-mapping', methods=['POST'])
def column_mapping():
    """Handle column mapping for device registration."""
    logger.info("="*80)
    logger.info("COLUMN MAPPING REQUEST")
    logger.info("="*80)
    
    selected_sheet = request.form.get('selected_sheet')
    logger.info(f"Selected sheet from form: {selected_sheet}")
    
    if not selected_sheet:
        logger.warning("No sheet selected")
        flash('Bitte wählen Sie ein Sheet aus', 'danger')
        return redirect(url_for('index'))
    
    # Store selected sheet in session
    session['selected_sheet'] = selected_sheet
    logger.info(f"Stored selected_sheet in session: {selected_sheet}")
    
    # Get the parsed data file from session
    parsed_data_file = session.get('parsed_data_file', '')
    original_filename = session.get('original_filename', '')
    logger.info(f"Parsed data file from session: {parsed_data_file}")
    
    if not parsed_data_file or not os.path.exists(parsed_data_file):
        logger.error("Parsed data file not found")
        flash('Keine Daten gefunden. Bitte laden Sie die Datei erneut hoch.', 'danger')
        return redirect(url_for('index'))
    
    # Read parsed data from file
    logger.info("Reading parsed data from file")
    with open(parsed_data_file, 'r') as f:
        parsed_data = json.load(f)
    
    logger.info(f"Available sheets in parsed data: {list(parsed_data.keys())}")
    
    # Get the selected sheet data
    if selected_sheet not in parsed_data:
        logger.error(f"Selected sheet '{selected_sheet}' not found in parsed data")
        flash('Sheet nicht gefunden', 'danger')
        return redirect(url_for('index'))
    
    # Get column names and preview
    df = pd.DataFrame(parsed_data[selected_sheet])
    columns = list(df.columns)
    logger.info(f"Sheet columns: {columns}")
    logger.info(f"Sheet has {len(df)} rows")
    
    # Check if application_id column exists (case-insensitive)
    has_application_id_column = any(col.lower() in ['application_id', 'app_id', 'applicationid'] for col in columns)
    logger.info(f"Has application_id column: {has_application_id_column}")
    
    # Generate preview HTML
    preview_html = df.head(5).to_html(
        classes='table is-bordered is-striped is-hoverable is-fullwidth',
        index=False,
        na_rep='N/A'
    )
    
    logger.info("Rendering column_mapping.html template")
    
    # Compute column statistics for validation
    column_stats = {}
    for col in columns:
        non_null_values = df[col].dropna()
        sample_values = non_null_values.head(3).astype(str).tolist()
        empty_count = df[col].isna().sum()
        non_empty_count = len(non_null_values)
        
        # Check if values look like hex keys (32 or 16 chars of hex)
        looks_like_key = False
        if sample_values:
            first_val = str(sample_values[0]).strip()
            if len(first_val) in [32, 16, 64]:  # Common key lengths
                looks_like_key = all(all(c in '0123456789ABCDEFabcdef' for c in str(v).strip() if c) for v in sample_values if v)
        
        column_stats[col] = {
            'samples': sample_values,
            'empty_count': int(empty_count),
            'non_empty_count': int(non_empty_count),
            'total_count': len(df),
            'empty_percent': round(empty_count / len(df) * 100, 1) if len(df) > 0 else 0,
            'looks_like_key': looks_like_key
        }
    
    logger.info("Column statistics computed")
    return render_template('column_mapping.html',
                         filename=original_filename,
                         selected_sheet=selected_sheet,
                         columns=columns,
                         row_count=len(df),
                         preview_html=preview_html,
                         has_application_id_column=has_application_id_column,
                         column_stats=json.dumps(column_stats))


@app.route('/process-mapping', methods=['POST'])
def process_mapping():
    """Process the column mapping and prepare for device registration."""
    logger.info("="*80)
    logger.info("PROCESS MAPPING REQUEST")
    logger.info("="*80)
    
    # Get column mappings from form
    column_mapping = {
        'dev_eui': request.form.get('dev_eui'),
        'name': request.form.get('name'),
        'application_id': request.form.get('application_id'),
        'device_profile_id': request.form.get('device_profile_id'),
        'nwk_key': request.form.get('nwk_key'),
        'app_key': request.form.get('app_key', ''),  # Optional
        'description': request.form.get('description', '')  # Optional
    }
    
    # Get manual application_id if no column is selected
    manual_application_id = request.form.get('manual_application_id', '').strip()
    if manual_application_id:
        column_mapping['manual_application_id'] = manual_application_id
        logger.info(f"Manual application_id provided: {manual_application_id}")
    
    # Get tag columns from form (user selected tags)
    tag_columns = request.form.getlist('tag_columns')
    logger.info(f"Tag columns selected: {tag_columns}")
    
    if tag_columns:
        column_mapping['tags'] = tag_columns
    else:
        column_mapping['tags'] = []
    
    logger.info(f"Column mapping received: {column_mapping}")
    
    # Validate required fields (application_id can come from column or manual input)
    required_fields = ['dev_eui', 'name', 'device_profile_id', 'nwk_key']
    missing_fields = [field for field in required_fields if not column_mapping[field]]
    
    # Check application_id: either from column or manual input
    if not column_mapping['application_id'] and not manual_application_id:
        missing_fields.append('application_id')
    
    if missing_fields:
        logger.error(f"Missing required fields: {missing_fields}")
        flash(f'Bitte ordnen Sie alle erforderlichen Felder zu: {", ".join(missing_fields)}', 'danger')
        return redirect(url_for('column_mapping'))
    
    # Validate column selections to catch common mistakes
    nwk_key_col = column_mapping['nwk_key'].lower() if column_mapping['nwk_key'] else ''
    app_key_col = column_mapping['app_key'].lower() if column_mapping['app_key'] else ''
    
    # Warning: If nwk_key selected is a SESSION key, not ROOT key
    if 'nwkskey' in nwk_key_col or 'appskey' in nwk_key_col:
        logger.warning(f"Potential issue: nwk_key column '{column_mapping['nwk_key']}' looks like a SESSION key, not ROOT key")
        flash('Achtung: Die Spalte für "Network Key" scheint ein Sitzungsschlüssel zu sein, nicht der Wurzelschlüssel. '
              'Bitte überprüfen Sie die Spaltenauswahl. Für 1.1.x-Geräte benötigen Sie den Netzwerk-Wurzelschlüssel.', 'warning')
    
    # Warning: If app_key selected is a SESSION key, not APP root key
    if app_key_col and ('nwkskey' in app_key_col):
        logger.warning(f"Potential issue: app_key column '{column_mapping['app_key']}' looks like the NETWORK Session Key, not the Application Key")
        flash('⚠️ KRITISCH: Die Spalte für "Application Key" ist auf "lora_nwkskey" eingestellt - das ist der NETWORK SESSION Key, nicht der Application Key! '
              'Für OTAA: Verwenden Sie "OTAA keys" oder "lora_appkey*". Für ABP: Es wird kein Application Key benötigt.', 'danger')
    
    # Warning: If app_key is empty but there are columns that look like app keys
    if not app_key_col:
        logger.warning(f"app_key column not selected - will not set Application Key in ChirpStack")
        flash('Informationen: Keine Spalte für "Application Key" ausgewählt. Stellen Sie sicher, dass dies beabsichtigt ist.', 'info')
    # Warning: If both nwk_key and app_key seem to be selected the same
    elif nwk_key_col == app_key_col:
        logger.warning(f"Potential issue: nwk_key and app_key are both set to same column '{column_mapping['nwk_key']}'")
        flash('Achtung: "Network Key" und "Application Key" sind auf die gleiche Spalte eingestellt. '
              'Bitte überprüfen Sie die Spaltenauswahl.', 'warning')
    
    # Store mapping in session
    session['column_mapping'] = column_mapping
    logger.info("Column mapping stored in session")
    
    # Redirect to registration preview page
    logger.info("Redirecting to registration preview")
    return redirect(url_for('registration_preview'))


@app.route('/registration-preview')
def registration_preview():
    """Show preview of devices to be registered."""
    logger.info("="*80)
    logger.info("REGISTRATION PREVIEW REQUEST")
    logger.info("="*80)
    
    # Get all required data from session
    parsed_data_file = session.get('parsed_data_file', '')
    selected_sheet = session.get('selected_sheet', '')
    column_mapping = session.get('column_mapping', {})
    original_filename = session.get('original_filename', '')
    
    logger.info(f"Selected sheet: {selected_sheet}")
    logger.info(f"Column mapping: {column_mapping}")
    
    # Validate we have all necessary data
    if not parsed_data_file or not selected_sheet or not column_mapping:
        logger.error("Missing required session data")
        flash('Session-Daten fehlen. Bitte starten Sie den Prozess erneut.', 'danger')
        return redirect(url_for('index'))
    
    if not os.path.exists(parsed_data_file):
        logger.error(f"Parsed data file not found: {parsed_data_file}")
        flash('Datei nicht gefunden. Bitte laden Sie die Datei erneut hoch.', 'danger')
        return redirect(url_for('index'))
    
    # Read parsed data
    with open(parsed_data_file, 'r') as f:
        parsed_data = json.load(f)
    
    if selected_sheet not in parsed_data:
        logger.error(f"Sheet '{selected_sheet}' not found in parsed data")
        flash('Sheet nicht gefunden.', 'danger')
        return redirect(url_for('index'))
    
    # Load data into DataFrame
    df = pd.DataFrame(parsed_data[selected_sheet])
    logger.info(f"Loaded {len(df)} devices from sheet")
    
    # Fetch LoRaWAN version information for the device profile
    lorawan_version_info = None
    try:
        sample_device_profile_id = str(df.iloc[0][column_mapping['device_profile_id']]).strip() if len(df) > 0 else None
        if sample_device_profile_id:
            from grpc_client import ChirpStackClient
            temp_client = ChirpStackClient(SERVER_URL, API_CODE)
            lorawan_version_info = temp_client.get_lorawan_version_from_profile_id(sample_device_profile_id, TENANT_ID)
            temp_client.close()
            logger.info(f"[Preview] Detected LoRaWAN version: {lorawan_version_info}")
    except Exception as e:
        logger.warning(f"[Preview] Could not fetch LoRaWAN version: {e}")
    
    # Map columns to device fields
    mapped_devices = []
    for idx, row in df.iterrows():
        # Handle application_id: use manual input if available, otherwise use column
        app_id = ''
        if column_mapping.get('manual_application_id'):
            app_id = column_mapping['manual_application_id']
        elif column_mapping['application_id']:
            app_id = str(row[column_mapping['application_id']])
        
        # Check if device is OTAA by looking for lora_joinmode column
        is_otaa = False
        if 'lora_joinmode' in df.columns:
            join_mode = str(row['lora_joinmode']).strip().upper() if pd.notna(row['lora_joinmode']) else 'ABP'
            is_otaa = join_mode == 'OTAA'
        
        # For OTAA devices, use OTAA keys column if available
        nwk_key_value = str(row[column_mapping['nwk_key']]) if column_mapping['nwk_key'] else ''
        if is_otaa and 'OTAA keys' in df.columns and pd.notna(row['OTAA keys']):
            # Override: for OTAA 1.0.x, use OTAA keys for the nwk_key field
            otaa_keys = str(row['OTAA keys']).strip()
            if otaa_keys:  # Only override if OTAA keys has a value
                nwk_key_value = otaa_keys
                logger.info(f"[Preview] Device {str(row[column_mapping['dev_eui']])}: OTAA detected, using 'OTAA keys' column: {nwk_key_value}")
        
        device = {
            'dev_eui': str(row[column_mapping['dev_eui']]) if column_mapping['dev_eui'] else '',
            'name': str(row[column_mapping['name']]) if column_mapping['name'] else '',
            'application_id': app_id,
            'device_profile_id': str(row[column_mapping['device_profile_id']]) if column_mapping['device_profile_id'] else '',
            'nwk_key': nwk_key_value,
            'app_key': str(row[column_mapping['app_key']]) if column_mapping.get('app_key') and column_mapping['app_key'] else '',
            'description': str(row[column_mapping['description']]) if column_mapping.get('description') and column_mapping['description'] else '',
            'is_otaa': is_otaa,
            'lorawan_version': lorawan_version_info  # NEW: Include version info
        }
        
        # DEBUG: Log the extraction
        if column_mapping.get('app_key'):
            logger.info(f"[Preview] Device {device['dev_eui']}: app_key column='{column_mapping['app_key']}', extracted value='{device['app_key']}', is_otaa={is_otaa}")
        
        # Extract tags
        tags = {}
        if column_mapping.get('tags'):
            for tag_col in column_mapping['tags']:
                tag_value = str(row[tag_col]).strip() if tag_col in row and pd.notna(row[tag_col]) else ''
                if tag_value:  # Only add non-empty tags
                    tags[tag_col] = tag_value
        
        device['tags'] = tags
        mapped_devices.append(device)
    
    logger.info(f"Mapped {len(mapped_devices)} devices successfully")
    
    # Validate mapped data for common issues
    data_audit = {
        'warnings': [],
        'statistics': {
            'total_devices': len(mapped_devices),
            'devices_with_empty_keys': 0,
            'devices_with_invalid_eui': 0,
            'devices_with_invalid_keys': 0,
            'devices_with_invalid_profile_id': 0
        },
        'unique_profile_ids': set(),
        'unique_app_ids': set()
    }
    
    # Check each device for issues
    for device in mapped_devices:
        # Track unique profile and app IDs
        data_audit['unique_profile_ids'].add(device['device_profile_id'])
        data_audit['unique_app_ids'].add(device['application_id'])
        
        # Check for empty network key
        if not device['nwk_key'] or str(device['nwk_key']).strip().upper() == 'NAN':
            data_audit['statistics']['devices_with_empty_keys'] += 1
        
        # Check DevEUI format (should be 16 hex chars)
        dev_eui = str(device['dev_eui']).strip()
        if not dev_eui or len(dev_eui) != 16 or not all(c in '0123456789ABCDEFabcdef' for c in dev_eui):
            data_audit['statistics']['devices_with_invalid_eui'] += 1
        
        # Check device_profile_id format (should be valid UUID)
        profile_id = str(device['device_profile_id']).strip()
        if not profile_id or not _is_valid_uuid(profile_id):
            data_audit['statistics']['devices_with_invalid_profile_id'] += 1
        
        # Check key formats (should be 32 hex chars)
        nwk_key = str(device['nwk_key']).strip()
        if nwk_key and (len(nwk_key) != 32 or not all(c in '0123456789ABCDEFabcdef' for c in nwk_key)):
            data_audit['statistics']['devices_with_invalid_keys'] += 1
        
        if device['app_key']:
            app_key = str(device['app_key']).strip()
            if app_key and (len(app_key) != 32 or not all(c in '0123456789ABCDEFabcdef' for c in app_key)):
                data_audit['statistics']['devices_with_invalid_keys'] += 1
    
    # Generate warnings based on audit
    if data_audit['statistics']['devices_with_empty_keys'] > 0:
        data_audit['warnings'].append(
            f"⚠️ {data_audit['statistics']['devices_with_empty_keys']} Gerät(e) haben keinen Network Key"
        )
    if data_audit['statistics']['devices_with_invalid_eui'] > 0:
        data_audit['warnings'].append(
            f"⚠️ {data_audit['statistics']['devices_with_invalid_eui']} Gerät(e) haben ungültige Device EUI (sollte 16 Hex-Zeichen sein)"
        )
    if data_audit['statistics']['devices_with_invalid_keys'] > 0:
        data_audit['warnings'].append(
            f"⚠️ {data_audit['statistics']['devices_with_invalid_keys']} Gerät(e) haben ungültige Schlüssel (sollten 32 Hex-Zeichen sein)"
        )
    if data_audit['statistics']['devices_with_invalid_profile_id'] > 0:
        data_audit['warnings'].append(
            f"⚠️ {data_audit['statistics']['devices_with_invalid_profile_id']} Gerät(e) haben ungültige Device Profile ID (sollte eine gültige UUID sein)"
        )
    
    # Warning if all devices use the same profile ID
    if len(data_audit['unique_profile_ids']) == 1:
        profile_id_str = list(data_audit['unique_profile_ids'])[0]
        data_audit['warnings'].append(
            f"ℹ️ Alle Geräte verwenden die gleiche Device Profile ID: <code>{profile_id_str}</code><br/>"
            f"<small>Bitte stellen Sie sicher, dass diese ID in Ihrer ChirpStack-Instanz existiert (nicht gelöscht oder in anderem Tenant)</small>"
        )
    
    logger.info(f"Data audit: {data_audit}")
    
    # Flash warnings if there are issues
    for warning in data_audit['warnings']:
        flash(warning, 'warning')
    
    # Convert sets to lists for JSON serialization in templates
    data_audit['unique_profile_ids'] = list(data_audit['unique_profile_ids'])
    data_audit['unique_app_ids'] = list(data_audit['unique_app_ids'])
    
    # Create preview DataFrame
    preview_df = pd.DataFrame(mapped_devices)
    preview_html = preview_df.head(10).to_html(
        classes='table is-bordered is-striped is-hoverable is-fullwidth',
        index=False,
        na_rep='N/A'
    )
    
    # Check server configuration
    server_configured = bool(SERVER_URL and API_CODE and TENANT_ID)
    logger.info(f"Server configured: {server_configured}")
    
    return render_template('registration_preview.html',
                         filename=original_filename,
                         device_count=len(mapped_devices),
                         preview_html=preview_html,
                         server_configured=server_configured,
                         server_url=SERVER_URL,
                         tenant_id=TENANT_ID,
                         data_audit=data_audit,
                         mapped_devices_json=json.dumps(mapped_devices))


@app.route('/start-registration', methods=['POST'])
def start_registration():
    """Start registration process - shows progress page"""
    # Store duplicate action in session
    duplicate_action = request.form.get('duplicate_action', 'skip')
    session['duplicate_action'] = duplicate_action
    
    # Store custom tags in session
    custom_tags_json = request.form.get('custom_tags', '{}')
    try:
        custom_tags = json.loads(custom_tags_json)
    except json.JSONDecodeError:
        custom_tags = {}
    
    session['custom_tags'] = custom_tags
    logger.info(f"Custom tags set: {custom_tags}")
    
    return render_template('registration_progress.html', duplicate_action=duplicate_action)


@app.route('/register-devices-stream', methods=['POST'])
def register_devices_stream():
    """Stream device registration progress in real-time using SSE."""
    
    def generate():
        """Generator function for Server-Sent Events"""
        try:
            # Check if server is configured
            logger.info(f"Registration stream started. SERVER_URL={SERVER_URL}, API_CODE={'SET' if API_CODE else 'NOT SET'}, TENANT_ID={TENANT_ID}")
            
            if not SERVER_URL or not API_CODE or not TENANT_ID:
                error_msg = 'Server nicht konfiguriert! Bitte gehen Sie zu "Einstellungen" und konfigurieren Sie:\n'
                if not SERVER_URL:
                    error_msg += '- SERVER_URL (z.B. localhost:8080)\n'
                if not API_CODE:
                    error_msg += '- API_CODE (ChirpStack API Token)\n'
                if not TENANT_ID:
                    error_msg += '- TENANT_ID\n'
                logger.error(f"Server configuration missing: {error_msg}")
                yield f"data: {json.dumps({'error': error_msg})}\n\n"
                return
            
            # Get duplicate action from session (set in start_registration)
            duplicate_action = session.get('duplicate_action', 'skip')
            logger.info(f"Streaming registration with duplicate_action: {duplicate_action}")
            
            # Get data from session
            parsed_data_file = session.get('parsed_data_file', '')
            selected_sheet = session.get('selected_sheet', '')
            column_mapping = session.get('column_mapping', {})
            
            if not parsed_data_file or not os.path.exists(parsed_data_file):
                yield f"data: {json.dumps({'error': 'Session data missing'})}\n\n"
                return
            
            # Read parsed data
            with open(parsed_data_file, 'r') as f:
                parsed_data = json.load(f)
            
            df = pd.DataFrame(parsed_data[selected_sheet])
            logger.info(f"Loaded {len(df)} devices from sheet")
            
            # Fetch LoRaWAN version information for the device profile
            # This will be used to determine correct key mapping
            lorawan_version_info = None
            try:
                # Get first device to extract device_profile_id
                sample_device_profile_id = str(df.iloc[0][column_mapping['device_profile_id']]).strip() if len(df) > 0 else None
                if sample_device_profile_id:
                    from grpc_client import ChirpStackClient
                    temp_client = ChirpStackClient(SERVER_URL, API_CODE)
                    lorawan_version_info = temp_client.get_lorawan_version_from_profile_id(sample_device_profile_id, TENANT_ID)
                    temp_client.close()
                    
                    if lorawan_version_info:
                        logger.info(f"[Registration] Detected LoRaWAN version: {lorawan_version_info['version']} (Profile: {lorawan_version_info['name']})")
                        yield f"data: {json.dumps({'status': 'info', 'message': f\"LoRaWAN-Version erkannt: {lorawan_version_info['version']}\", 'version': lorawan_version_info['version']})}\n\n"
                    else:
                        logger.warning(f"[Registration] Could not determine LoRaWAN version for profile {sample_device_profile_id}")
            except Exception as e:
                logger.warning(f"[Registration] Error fetching LoRaWAN version: {e}")
                # Continue anyway - we have fallback logic
            
            custom_tags = session.get('custom_tags', {})
            logger.info(f"Custom tags from session: {custom_tags}")
            
            # Map columns to device fields
            devices_to_register = []
            for idx, row in df.iterrows():
                # Handle application_id: use manual input if available, otherwise use column
                app_id = ''
                if column_mapping.get('manual_application_id'):
                    app_id = column_mapping['manual_application_id']
                elif column_mapping['application_id']:
                    app_id = str(row[column_mapping['application_id']]).strip()
                
                # Check if device is OTAA by looking for lora_joinmode column
                is_otaa = False
                if 'lora_joinmode' in df.columns:
                    join_mode = str(row['lora_joinmode']).strip().upper() if pd.notna(row['lora_joinmode']) else 'ABP'
                    is_otaa = join_mode == 'OTAA'
                
                # For OTAA devices, use OTAA keys column if available
                nwk_key_value = str(row[column_mapping['nwk_key']]).strip()
                if is_otaa and 'OTAA keys' in df.columns and pd.notna(row['OTAA keys']):
                    # Override: for OTAA 1.0.x, use OTAA keys for the nwk_key field
                    otaa_keys = str(row['OTAA keys']).strip()
                    if otaa_keys:  # Only override if OTAA keys has a value
                        nwk_key_value = otaa_keys
                        logger.info(f"[Registration] Device {str(row[column_mapping['dev_eui']]).strip()}: OTAA detected, using 'OTAA keys' column for nwk_key field: {nwk_key_value}")
                
                device = {
                    'dev_eui': str(row[column_mapping['dev_eui']]).strip(),
                    'name': str(row[column_mapping['name']]).strip(),
                    'application_id': app_id,
                    'device_profile_id': str(row[column_mapping['device_profile_id']]).strip(),
                    'nwk_key': nwk_key_value,
                    'app_key': str(row[column_mapping['app_key']]).strip() if column_mapping.get('app_key') and column_mapping['app_key'] else '',
                    'description': str(row[column_mapping['description']]).strip() if column_mapping.get('description') and column_mapping['description'] else '',
                    'is_otaa': is_otaa,
                    'lorawan_version': lorawan_version_info  # NEW: Pass version info
                }
                
                # DEBUG: Log the extraction
                if column_mapping.get('app_key'):
                    logger.info(f"[Registration] Device {device['dev_eui']}: app_key column='{column_mapping['app_key']}', extracted value='{device['app_key']}', is_otaa={is_otaa}")
                
                # Extract tags from columns
                tags = {}
                if column_mapping.get('tags'):
                    for tag_col in column_mapping['tags']:
                        tag_value = str(row[tag_col]).strip() if tag_col in row and pd.notna(row[tag_col]) else ''
                        if tag_value:  # Only add non-empty tags
                            tags[tag_col] = tag_value
                
                # Add custom tags (these are user-defined and apply to all devices)
                if custom_tags:
                    tags.update(custom_tags)
                    logger.info(f"Device {device['dev_eui']} tags after custom merge: {tags}")
                
                device['tags'] = tags
                devices_to_register.append(device)
            
            total = len(devices_to_register)
            
            # Send initial status
            yield f"data: {json.dumps({'status': 'starting', 'total': total, 'current': 0})}\n\n"
            
            logger.info(f"Starting parallel device registration for {total} devices")
            
            results = {'successful': [], 'failed': []}
            results_lock = threading.Lock()  # Thread-safe access to results
            completed_count = [0]  # Use list to allow mutation in nested function
            
            # Define worker function for parallel processing
            def register_single_device(idx_device_tuple):
                """Register a single device - worker function for thread pool"""
                idx, device = idx_device_tuple
                try:
                    # Create client for this thread
                    from grpc_client import ChirpStackClient
                    thread_client = ChirpStackClient(SERVER_URL, API_CODE)
                    
                    logger.info(f"[Worker-{idx}] === STARTING DEVICE REGISTRATION ===")
                    logger.info(f"[Worker-{idx}] Device: {device['dev_eui']} ({device.get('name', 'NO_NAME')})")
                    logger.info(f"[Worker-{idx}] Application ID: {device.get('application_id', 'NO_APP_ID')}")
                    logger.info(f"[Worker-{idx}] Device Profile ID: {device.get('device_profile_id', 'NO_PROFILE_ID')}")
                    
                    connected, conn_msg = thread_client.connect()
                    logger.info(f"[Worker-{idx}] Connection result: connected={connected}, msg={conn_msg}")
                    
                    if not connected:
                        logger.error(f"[Worker-{idx}] Connection failed")
                        return {
                            'idx': idx,
                            'device': device,
                            'result': 'failed',
                            'message': f'Connection failed: {conn_msg}'
                        }
                    
                    # Check if device exists
                    device_exists = thread_client.device_exists(device['dev_eui'])
                    logger.info(f"[Worker-{idx}] Device {device['dev_eui']}: exists={device_exists}, action={duplicate_action}")
                    
                    if device_exists and duplicate_action == 'skip':
                        logger.info(f"[Worker-{idx}] Device exists and action is skip - adding to failed list")
                        with results_lock:
                            results['failed'].append({
                                'dev_eui': device['dev_eui'],
                                'name': device['name'],
                                'error': 'Gerät existiert bereits (übersprungen)'
                            })
                        return {
                            'idx': idx,
                            'device': device,
                            'result': 'skipped',
                            'message': 'Bereits vorhanden'
                        }
                    
                    if device_exists and duplicate_action == 'replace':
                        logger.info(f"[Worker-{idx}] Device exists and action is replace - deleting device")
                        deleted, del_msg = thread_client.delete_device(device['dev_eui'])
                        if not deleted:
                            logger.error(f"[Worker-{idx}] Failed to delete: {del_msg}")
                            with results_lock:
                                results['failed'].append({
                                    'dev_eui': device['dev_eui'],
                                    'name': device['name'],
                                    'error': f'Fehler beim Löschen: {del_msg}'
                                })
                            return {
                                'idx': idx,
                                'device': device,
                                'result': 'failed',
                                'message': 'Löschen fehlgeschlagen'
                            }
                        logger.info(f"[Worker-{idx}] Device {device['dev_eui']} deleted successfully")
                    
                    # Create device
                    logger.info(f"[Worker-{idx}] CALLING create_device...")
                    device_created, create_msg = thread_client.create_device(
                        dev_eui=device['dev_eui'],
                        name=device['name'],
                        application_id=device['application_id'],
                        device_profile_id=device['device_profile_id'],
                        description=device['description'],
                        tags=device.get('tags', {}) if device.get('tags') else None
                    )
                    logger.info(f"[Worker-{idx}] create_device returned: created={device_created}, msg={create_msg}")
                    
                    if not device_created:
                        logger.error(f"[Worker-{idx}] Device creation failed: {create_msg}")
                        with results_lock:
                            results['failed'].append({
                                'dev_eui': device['dev_eui'],
                                'name': device['name'],
                                'error': create_msg
                            })
                        logger.info(f"[Worker-{idx}] Device added to FAILED list (total failed now: {len(results['failed'])})")
                        return {
                            'idx': idx,
                            'device': device,
                            'result': 'failed',
                            'message': create_msg[:50]
                        }
                    
                    # Set device keys
                    logger.info(f"[Worker-{idx}] CALLING create_device_keys...")
                    keys_set, keys_msg = thread_client.create_device_keys(
                        dev_eui=device['dev_eui'],
                        nwk_key=device['nwk_key'],
                        app_key=device['app_key'] if device['app_key'] else None,
                        lorawan_version=device.get('lorawan_version')  # NEW: Use actual version
                    )
                    logger.info(f"[Worker-{idx}] create_device_keys returned: set={keys_set}, msg={keys_msg}")
                    
                    thread_client.close()
                    
                    if not keys_set:
                        logger.warning(f"[Worker-{idx}] Keys not set but device was created - adding to successful (with warning)")
                        with results_lock:
                            results['successful'].append({
                                'dev_eui': device['dev_eui'],
                                'name': device['name'],
                                'warning': f'Device created but keys not set: {keys_msg}'
                            })
                        logger.info(f"[Worker-{idx}] Device added to SUCCESSFUL list (total successful now: {len(results['successful'])})")
                        return {
                            'idx': idx,
                            'device': device,
                            'result': 'warning',
                            'message': 'Keys nicht gesetzt'
                        }
                    
                    logger.info(f"[Worker-{idx}] SUCCESS - Device fully created and keys set, adding to successful list")
                    with results_lock:
                        results['successful'].append({
                            'dev_eui': device['dev_eui'],
                            'name': device['name']
                        })
                    logger.info(f"[Worker-{idx}] Device added to SUCCESSFUL list (total successful now: {len(results['successful'])})")
                    
                    return {
                        'idx': idx,
                        'device': device,
                        'result': 'success',
                        'message': 'Erfolgreich'
                    }
                
                except Exception as e:
                    logger.error(f"[Worker-{idx}] EXCEPTION occurred: {str(e)}", exc_info=True)
                    with results_lock:
                        results['failed'].append({
                            'dev_eui': device.get('dev_eui', 'N/A'),
                            'name': device.get('name', 'N/A'),
                            'error': str(e)
                        })
                    logger.info(f"[Worker-{idx}] Device added to FAILED list due to exception (total failed now: {len(results['failed'])})")
                    return {
                        'idx': idx,
                        'device': device,
                        'result': 'failed',
                        'message': str(e)[:50]
                    }
            
            # Use ThreadPoolExecutor for parallel processing (10 worker threads)
            num_workers = min(10, max(1, total // 20))  # Scale workers based on device count
            logger.info(f"Starting parallel registration with {num_workers} workers for {total} devices")
            
            with ThreadPoolExecutor(max_workers=num_workers) as executor:
                # Map the worker function to all devices
                futures = {executor.submit(register_single_device, (idx + 1, device)): idx 
                          for idx, device in enumerate(devices_to_register)}
                
                # Process results as they complete
                for future in as_completed(futures):
                    try:
                        result = future.result()
                        idx = result['idx']
                        device = result['device']
                        
                        completed_count[0] += 1
                        
                        yield f"data: {json.dumps({
                            'status': 'processing',
                            'current': completed_count[0],
                            'total': total,
                            'device': device['name'],
                            'dev_eui': device['dev_eui'],
                            'application_id': device.get('application_id', ''),
                            'device_profile_id': device.get('device_profile_id', ''),
                            'result': result['result'],
                            'message': result['message']
                        })}\n\n"
                        
                    except Exception as e:
                        logger.error(f"Error processing future: {str(e)}", exc_info=True)
                        completed_count[0] += 1
                        yield f"data: {json.dumps({
                            'status': 'processing',
                            'current': completed_count[0],
                            'total': total,
                            'result': 'failed',
                            'message': f'Worker error: {str(e)[:50]}'
                        })}\n\n"
            
            logger.info(f"="*80)
            logger.info(f"REGISTRATION COMPLETE - FINAL SUMMARY")
            logger.info(f"="*80)
            logger.info(f"Successful: {len(results['successful'])}")
            logger.info(f"Failed: {len(results['failed'])}")
            logger.info(f"Successful devices: {[d['dev_eui'] for d in results['successful']]}")
            logger.info(f"Failed devices: {[(d['dev_eui'], d.get('error', 'N/A')) for d in results['failed']]}")
            logger.info(f"="*80)
            
            # Store results in session
            session['registration_results'] = results
            
            # Send completion
            yield f"data: {json.dumps({
                'status': 'complete',
                'successful': len(results['successful']),
                'failed': len(results['failed'])
            })}\n\n"
            
            
        except Exception as e:
            logger.error(f"Streaming error: {e}", exc_info=True)
            yield f"data: {json.dumps({'error': str(e)})}\n\n"
    
    return Response(stream_with_context(generate()), mimetype='text/event-stream')


@app.route('/register-devices', methods=['POST'])
def register_devices():
    """Execute device registration via gRPC."""
    logger.info("="*80)
    logger.info("REGISTER DEVICES REQUEST - STARTING")
    logger.info("="*80)
    
    # Check server configuration
    if not SERVER_URL or not API_CODE or not TENANT_ID:
        logger.error("Server not configured")
        flash('Server ist nicht konfiguriert. Bitte konfigurieren Sie zuerst die Server-Verbindung.', 'danger')
        return redirect(url_for('server_config'))
    
    # Get all required data from session
    parsed_data_file = session.get('parsed_data_file', '')
    selected_sheet = session.get('selected_sheet', '')
    column_mapping = session.get('column_mapping', {})
    
    if not parsed_data_file or not selected_sheet or not column_mapping:
        logger.error("Missing session data")
        flash('Session-Daten fehlen. Bitte starten Sie den Prozess erneut.', 'danger')
        return redirect(url_for('index'))
    
    if not os.path.exists(parsed_data_file):
        logger.error(f"Parsed data file not found: {parsed_data_file}")
        flash('Datei nicht gefunden.', 'danger')
        return redirect(url_for('index'))
    
    # Read parsed data
    with open(parsed_data_file, 'r') as f:
        parsed_data = json.load(f)
    
    df = pd.DataFrame(parsed_data[selected_sheet])
    logger.info(f"Starting registration for {len(df)} devices")
    
    # Get duplicate handling action
    duplicate_action = request.form.get('duplicate_action', 'skip')  # 'skip' or 'replace'
    logger.info(f"Duplicate action: {duplicate_action}")
    
    # Map columns to device fields
    devices_to_register = []
    for idx, row in df.iterrows():
        # Handle application_id: use manual input if available, otherwise use column
        app_id = ''
        if column_mapping.get('manual_application_id'):
            app_id = column_mapping['manual_application_id']
        elif column_mapping['application_id']:
            app_id = str(row[column_mapping['application_id']]).strip()
        
        device = {
            'dev_eui': str(row[column_mapping['dev_eui']]).strip(),
            'name': str(row[column_mapping['name']]).strip(),
            'application_id': app_id,
            'device_profile_id': str(row[column_mapping['device_profile_id']]).strip(),
            'nwk_key': str(row[column_mapping['nwk_key']]).strip(),
            'app_key': str(row[column_mapping['app_key']]).strip() if column_mapping.get('app_key') and column_mapping['app_key'] else '',
            'description': str(row[column_mapping['description']]).strip() if column_mapping.get('description') and column_mapping['description'] else ''
        }
        devices_to_register.append(device)
    
    # Initialize results tracking
    results = {
        'total': len(devices_to_register),
        'successful': [],
        'failed': []
    }
    
    # Import gRPC client
    try:
        from grpc_client import ChirpStackClient
        logger.info("Imported ChirpStackClient successfully")
    except ImportError as e:
        logger.error(f"Failed to import ChirpStackClient: {e}")
        flash('Fehler beim Laden des gRPC Clients.', 'danger')
        return redirect(url_for('registration_preview'))
    
    # Create gRPC client
    try:
        logger.info(f"Creating ChirpStack client for {SERVER_URL}")
        client = ChirpStackClient(SERVER_URL, API_CODE)
        logger.info("ChirpStack client created successfully")
        
        # Connect to ChirpStack
        connected, conn_msg = client.connect()
        if not connected:
            logger.error(f"Failed to connect to ChirpStack: {conn_msg}")
            flash(f'Fehler beim Verbinden mit ChirpStack: {conn_msg}', 'danger')
            return redirect(url_for('registration_preview'))
        logger.info(f"Connected to ChirpStack: {conn_msg}")
        
    except Exception as e:
        logger.error(f"Failed to create ChirpStack client: {e}", exc_info=True)
        flash(f'Fehler beim Verbinden mit ChirpStack: {str(e)}', 'danger')
        return redirect(url_for('registration_preview'))
    
    # Register each device
    for idx, device in enumerate(devices_to_register, 1):
        logger.info(f"Registering device {idx}/{len(devices_to_register)}: {device['name']} ({device['dev_eui']})")
        
        try:
            # Check if device already exists
            device_exists = client.device_exists(device['dev_eui'])
            
            if device_exists:
                logger.info(f"Device {device['dev_eui']} already exists")
                
                if duplicate_action == 'skip':
                    # Skip this device
                    logger.warning(f"Skipping existing device {device['dev_eui']}")
                    results['failed'].append({
                        'dev_eui': device['dev_eui'],
                        'name': device['name'],
                        'error': 'Gerät existiert bereits (übersprungen)'
                    })
                    continue
                    
                elif duplicate_action == 'replace':
                    # Delete existing device first
                    logger.info(f"Deleting existing device {device['dev_eui']} for replacement")
                    deleted, del_msg = client.delete_device(device['dev_eui'])
                    if not deleted:
                        logger.error(f"Failed to delete existing device {device['dev_eui']}: {del_msg}")
                        results['failed'].append({
                            'dev_eui': device['dev_eui'],
                            'name': device['name'],
                            'error': f'Fehler beim Löschen des existierenden Geräts: {del_msg}'
                        })
                        continue
                    logger.info(f"Existing device {device['dev_eui']} deleted successfully")
            
            # Create device
            device_created, create_msg = client.create_device(
                dev_eui=device['dev_eui'],
                name=device['name'],
                application_id=device['application_id'],
                device_profile_id=device['device_profile_id'],
                description=device['description']
            )
            
            if not device_created:
                logger.error(f"Failed to create device {device['name']}: {create_msg}")
                results['failed'].append({
                    'dev_eui': device['dev_eui'],
                    'name': device['name'],
                    'error': create_msg
                })
                continue
            
            logger.info(f"Device {device['name']} created successfully: {create_msg}")
            
            # Set device keys
            keys_set, keys_msg = client.create_device_keys(
                dev_eui=device['dev_eui'],
                nwk_key=device['nwk_key'],
                app_key=device['app_key'] if device['app_key'] else None,
                is_otaa=device.get('is_otaa', True)
            )
            
            if not keys_set:
                logger.warning(f"Failed to set keys for device {device['name']}: {keys_msg}")
                results['successful'].append({
                    'dev_eui': device['dev_eui'],
                    'name': device['name'],
                    'warning': f'Device created but keys not set: {keys_msg}'
                })
            else:
                logger.info(f"Keys set successfully for device {device['name']}: {keys_msg}")
                results['successful'].append({
                    'dev_eui': device['dev_eui'],
                    'name': device['name']
                })
        
        except Exception as e:
            logger.error(f"Error registering device {device['name']}: {e}", exc_info=True)
            results['failed'].append({
                'dev_eui': device['dev_eui'],
                'name': device['name'],
                'error': str(e)
            })
    
    # Store results in session for display
    session['registration_results'] = results
    
    logger.info("="*80)
    logger.info(f"REGISTRATION COMPLETE - Success: {len(results['successful'])}, Failed: {len(results['failed'])}")
    logger.info("="*80)
    
    return redirect(url_for('registration_results'))


def generate_registration_report(results, server_info=None):
    """Generate an Excel report with registration results."""
    try:
        # Set defaults for server_info if not provided
        if server_info is None:
            server_info = {
                'server_url': session.get('server_url', 'N/A'),
                'api_code': session.get('api_code', 'N/A')[:20] + "..." if session.get('api_code') else 'N/A',
                'tenant_id': session.get('tenant_id', 'N/A')
            }
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Registration Report"
        
        # Define colors and styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        success_font = Font(color="006100", bold=True)
        failed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        failed_font = Font(color="9C0006", bold=True)
        info_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = "LoRaWAN Device Registration Report"
        cell.font = Font(bold=True, size=14)
        cell.alignment = center_alignment
        row += 2
        
        # Summary Section
        successful_count = len(results.get('successful', []))
        failed_count = len(results.get('failed', []))
        total_count = successful_count + failed_count
        
        ws[f'A{row}'].value = "Summary:"
        ws[f'A{row}'].font = info_font
        row += 1
        
        ws[f'A{row}'].value = "Total Devices"
        ws[f'B{row}'].value = total_count
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'].value = "Successful"
        ws[f'B{row}'].value = successful_count
        ws[f'B{row}'].font = success_font
        ws[f'B{row}'].fill = success_fill
        row += 1
        
        ws[f'A{row}'].value = "Failed"
        ws[f'B{row}'].value = failed_count
        ws[f'B{row}'].font = failed_font
        ws[f'B{row}'].fill = failed_fill
        row += 2
        
        # Device Details Table Header
        headers = ["DevEUI", "Device Name", "Status", "Details", "Application ID", "Device Profile ID"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        row += 1
        
        # Add successful devices
        for device in results.get('successful', []):
            ws.cell(row=row, column=1).value = device.get('dev_eui', 'N/A')
            ws.cell(row=row, column=2).value = device.get('name', 'N/A')
            
            status_cell = ws.cell(row=row, column=3)
            status_cell.value = "✓ SUCCESS"
            status_cell.fill = success_fill
            status_cell.font = success_font
            status_cell.alignment = center_alignment
            
            details = device.get('warning', '')
            ws.cell(row=row, column=4).value = details if details else "Device created and keys set"
            
            ws.cell(row=row, column=5).value = device.get('application_id', 'N/A')
            ws.cell(row=row, column=6).value = device.get('device_profile_id', 'N/A')
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = border
            
            row += 1
        
        # Add failed devices
        for device in results.get('failed', []):
            ws.cell(row=row, column=1).value = device.get('dev_eui', 'N/A')
            ws.cell(row=row, column=2).value = device.get('name', 'N/A')
            
            status_cell = ws.cell(row=row, column=3)
            status_cell.value = "✗ FAILED"
            status_cell.fill = failed_fill
            status_cell.font = failed_font
            status_cell.alignment = center_alignment
            
            error_msg = device.get('error', 'Unknown error')
            ws.cell(row=row, column=4).value = error_msg
            
            ws.cell(row=row, column=5).value = "N/A"
            ws.cell(row=row, column=6).value = "N/A"
            
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = border
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info("Excel report generated successfully")
        return output
    
    except Exception as e:
        logger.error(f"Error generating Excel report: {e}", exc_info=True)
        return None


@app.route('/download-report', methods=['GET', 'POST'])
def download_report():
    """Download registration results as Excel file."""
    try:
        # Try to get results from POST data first, then session
        results = None
        server_info = {}
        
        if request.method == 'POST':
            results_data = request.form.get('resultsData')
            if results_data:
                try:
                    data = json.loads(results_data)
                    results = {
                        'successful': data.get('successful', []),
                        'failed': data.get('failed', [])
                    }
                    server_info = {
                        'server_url': data.get('server_url', 'N/A'),
                        'api_code': data.get('api_code', 'N/A'),
                        'tenant_id': data.get('tenant_id', 'N/A')
                    }
                    logger.info("Results retrieved from POST data")
                except json.JSONDecodeError as e:
                    logger.error(f"Failed to parse results data: {e}")
        
        # Fall back to session if POST data not available
        if not results:
            results = session.get('registration_results', {})
            server_info = {
                'server_url': session.get('server_url', 'N/A'),
                'api_code': session.get('api_code', 'N/A')[:20] + "..." if session.get('api_code') else 'N/A',
                'tenant_id': session.get('tenant_id', 'N/A')
            }
            logger.info("Results retrieved from session")
        
        if not results:
            logger.warning("No registration results found for download")
            flash('Keine Registrierungsergebnisse zum Herunterladen gefunden.', 'error')
            return redirect(url_for('index'))
        
        excel_file = generate_registration_report(results, server_info)
        
        if excel_file is None:
            logger.error("Failed to generate Excel report")
            flash('Fehler beim Erstellen des Berichts.', 'error')
            return redirect(url_for('index'))
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"LoRaWAN_Registration_Report_{timestamp}.xlsx"
        
        logger.info(f"Sending Excel report: {filename}")
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        logger.error(f"Error downloading report: {e}", exc_info=True)
        flash('Fehler beim Herunterladen des Berichts.', 'error')
        return redirect(url_for('index'))


@app.route('/registration-results')
def registration_results():
    """Display registration results."""
    logger.info("="*80)
    logger.info("REGISTRATION RESULTS PAGE")
    logger.info("="*80)
    
    results = session.get('registration_results', {})
    
    if not results:
        logger.warning("No registration results found in session")
        flash('Keine Registrierungsergebnisse gefunden.', 'warning')
        return redirect(url_for('index'))
    
    logger.info(f"Displaying results: {results.get('total', 0)} total, {len(results.get('successful', []))} successful, {len(results.get('failed', []))} failed")
    
    return render_template('registration_results.html', 
                         results=results,
                         now=datetime.now())


@app.route('/change_sheet', methods=['POST'])
def change_sheet():
    """Handle sheet change request."""
    sheet_name = request.form.get('sheet_name')
    filepath = session.get('filepath')
    original_filename = session.get('original_filename')
    sheet_names = session.get('sheet_names', [])
    
    if not filepath or not os.path.exists(filepath):
        flash('Datei nicht gefunden. Bitte laden Sie die Datei erneut hoch.', 'danger')
        return redirect(url_for('index'))
    
    try:
        # Read selected sheet
        df = pd.read_excel(filepath, sheet_name=sheet_name)
        
        # Convert DataFrame to HTML table
        # Get first 100 rows for preview
        preview_df = df.head(100)
        
        # Convert to HTML with custom classes
        table_html = preview_df.to_html(
            classes='table is-bordered is-striped is-hoverable is-fullwidth',
            index=False,
            na_rep='N/A'
        )
        
        # Get file info
        file_info = {
            'filename': original_filename,
            'rows': len(df),
            'columns': len(df.columns),
            'column_names': list(df.columns),
            'current_sheet': sheet_name,
            'sheet_names': sheet_names
        }
        
        return render_template('preview.html', 
                             table_html=table_html, 
                             file_info=file_info)
    
    except Exception as e:
        flash(f'Fehler beim Lesen des Blattes: {str(e)}', 'danger')
        return redirect(url_for('index'))


@app.route('/cleanup', methods=['POST'])
def cleanup():
    """Clean up uploaded file and return to home."""
    filepath = session.get('filepath')
    if filepath and os.path.exists(filepath):
        os.remove(filepath)
    session.clear()
    return redirect(url_for('index'))


@app.errorhandler(413)
def too_large(e):
    """Handle file too large error."""
    flash('Datei ist zu groß. Maximale Größe beträgt 16MB', 'danger')
    return redirect(url_for('index'))


# ============================================================================
# DEVICE MANAGEMENT ROUTES
# ============================================================================

@app.route('/device-management')
def device_management():
    """Device management page - view and delete devices."""
    global SERVER_URL, API_CODE
    
    logger.info("="*80)
    logger.info("DEVICE MANAGEMENT PAGE")
    logger.info("="*80)
    
    # Check if server is configured
    if not SERVER_URL or not API_CODE:
        flash('Bitte konfigurieren Sie zuerst den Server', 'warning')
        return redirect(url_for('server_config'))
    
    return render_template('device_management.html',
                         server_url=SERVER_URL)


@app.route('/api/list-devices', methods=['POST'])
def api_list_devices():
    """API endpoint to list devices."""
    global SERVER_URL, API_CODE
    
    try:
        # Get parameters from request
        application_id = request.form.get('application_id', '').strip()
        search = request.form.get('search', '').strip()
        limit = int(request.form.get('limit', 1000))
        offset = int(request.form.get('offset', 0))
        
        logger.info(f"=== LIST DEVICES REQUEST ===")
        logger.info(f"Application ID: '{application_id}'")
        logger.info(f"Search: '{search}'")
        logger.info(f"Limit: {limit}, Offset: {offset}")
        logger.info(f"Server URL: {SERVER_URL}")
        logger.info(f"API Key configured: {bool(API_CODE)}")
        
        # Validate application_id is provided
        if not application_id:
            logger.warning("Application ID is required but was not provided")
            return {'success': False, 'message': 'Application ID ist erforderlich'}, 400
        
        # Create gRPC client
        from grpc_client import ChirpStackClient
        client = ChirpStackClient(SERVER_URL, API_CODE)
        logger.info(f"gRPC client created")
        
        # Connect
        logger.info(f"Attempting to connect to ChirpStack...")
        connected, conn_msg = client.connect()
        if not connected:
            logger.error(f"Connection failed: {conn_msg}")
            return {'success': False, 'message': f'Connection failed: {conn_msg}'}, 500
        logger.info(f"Connected successfully: {conn_msg}")
        
        # List devices
        logger.info(f"Calling list_devices with application_id='{application_id}'...")
        success, result = client.list_devices(
            application_id=application_id,
            limit=limit,
            offset=offset,
            search=search
        )
        
        client.close()
        
        if success:
            logger.info(f"✓ Successfully retrieved {len(result['devices'])} devices (total: {result['total_count']})")
            return {'success': True, 'data': result}
        else:
            logger.error(f"✗ Failed to list devices. Error: {result}")
            return {'success': False, 'message': f'ChirpStack error: {result}'}, 500
            
    except Exception as e:
        logger.error(f"✗ Exception in api_list_devices: {type(e).__name__}: {str(e)}", exc_info=True)
        return {'success': False, 'message': str(e)}, 500


@app.route('/api/update-device-tags', methods=['POST'])
def api_update_device_tags():
    """API endpoint to update device tags."""
    global SERVER_URL, API_CODE
    
    try:
        # Get parameters from request
        dev_eui = request.form.get('dev_eui', '').strip()
        tags_str = request.form.get('tags', '').strip()
        
        logger.info(f"=== UPDATE DEVICE TAGS REQUEST ===")
        logger.info(f"Device EUI: '{dev_eui}'")
        logger.info(f"Tags string: '{tags_str}'")
        
        # Validate dev_eui is provided
        if not dev_eui:
            logger.warning("Device EUI is required but was not provided")
            return {'success': False, 'message': 'Device EUI ist erforderlich'}, 400
        
        # Parse tags from string format: "key1:value1|key2:value2"
        tags_dict = {}
        if tags_str:
            try:
                tag_pairs = [pair.strip() for pair in tags_str.split('|') if pair.strip()]
                for pair in tag_pairs:
                    if ':' not in pair:
                        return {'success': False, 'message': f'Ungültiges Tag-Format: "{pair}". Format sollte sein: key:value'}, 400
                    key, value = pair.split(':', 1)
                    key = key.strip()
                    value = value.strip()
                    if not key:
                        return {'success': False, 'message': 'Tag-Schlüssel kann nicht leer sein'}, 400
                    tags_dict[key] = value
            except Exception as e:
                logger.error(f"Failed to parse tags: {e}")
                return {'success': False, 'message': f'Fehler beim Parsen von Tags: {str(e)}'}, 400
        
        # Create gRPC client
        from grpc_client import ChirpStackClient
        client = ChirpStackClient(SERVER_URL, API_CODE)
        logger.info(f"gRPC client created")
        
        # Connect
        logger.info(f"Attempting to connect to ChirpStack...")
        connected, conn_msg = client.connect()
        if not connected:
            logger.error(f"Connection failed: {conn_msg}")
            return {'success': False, 'message': f'Verbindung fehlgeschlagen: {conn_msg}'}, 500
        logger.info(f"Connected successfully: {conn_msg}")
        
        # Update device tags
        logger.info(f"Calling update_device with dev_eui='{dev_eui}', tags={tags_dict}...")
        success, message = client.update_device(dev_eui, tags=tags_dict)
        
        client.close()
        
        if success:
            logger.info(f"✓ Successfully updated tags for device {dev_eui}")
            return {'success': True, 'message': message}
        else:
            logger.error(f"✗ Failed to update device. Error: {message}")
            return {'success': False, 'message': message}, 500
            
    except Exception as e:
        logger.error(f"✗ Exception in api_update_device_tags: {type(e).__name__}: {str(e)}", exc_info=True)
        return {'success': False, 'message': str(e)}, 500


@app.route('/api/generate-selected-devices-report', methods=['POST'])
def api_generate_selected_devices_report():
    """Generate Excel report for selected existing devices."""
    global SERVER_URL, API_CODE
    
    try:
        # Get list of dev_euis from request
        dev_euis_str = request.form.get('dev_euis', '').strip()
        
        if not dev_euis_str:
            logger.warning("No devices selected for report generation")
            return {'success': False, 'message': 'Keine Geräte ausgewählt'}, 400
        
        # Parse dev_euis (comma-separated)
        dev_euis = [eui.strip() for eui in dev_euis_str.split(',') if eui.strip()]
        
        if not dev_euis:
            return {'success': False, 'message': 'Keine gültigen Geräte-EUIs'}, 400
        
        logger.info(f"=== GENERATE SELECTED DEVICES REPORT ===")
        logger.info(f"Requested devices: {len(dev_euis)}")
        
        # Create gRPC client
        from grpc_client import ChirpStackClient
        client = ChirpStackClient(SERVER_URL, API_CODE)
        
        # Connect
        connected, conn_msg = client.connect()
        if not connected:
            logger.error(f"Connection failed: {conn_msg}")
            return {'success': False, 'message': f'Verbindung fehlgeschlagen: {conn_msg}'}, 500
        
        # Fetch each device's full data
        devices = []
        for dev_eui in dev_euis:
            success, device_data = client.get_device(dev_eui)
            
            if success:
                # Format as registration result
                devices.append({
                    'dev_eui': device_data['dev_eui'],
                    'name': device_data['name'],
                    'status': 'success',
                    'details': 'Loaded from existing devices',
                    'application_id': device_data.get('application_id', 'N/A'),
                    'device_profile_id': device_data.get('device_profile_id', 'N/A'),
                    'description': device_data.get('description', ''),
                    'tags': device_data.get('tags', {})
                })
                logger.info(f"✓ Loaded device: {dev_eui}")
            else:
                logger.warning(f"✗ Failed to load device: {dev_eui}")
        
        client.close()
        
        if not devices:
            return {'success': False, 'message': 'Keine Geräte konnten geladen werden'}, 400
        
        logger.info(f"Successfully loaded {len(devices)} devices for report")
        
        # Format results for report generation
        results = {
            'successful': devices,
            'failed': []
        }
        
        # Generate report
        excel_file = generate_registration_report(results)
        
        if excel_file is None:
            logger.error("Failed to generate Excel report")
            return {'success': False, 'message': 'Fehler beim Erstellen des Berichts'}, 500
        
        # Stream the file as a response
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"LoRaWAN_Devices_Report_{timestamp}.xlsx"
        
        logger.info(f"Generated report: {filename}")
        
        # Read file content and return it via JSON response
        excel_file.seek(0)
        file_content = excel_file.read()
        
        # Encode as base64 for transfer
        import base64
        encoded_file = base64.b64encode(file_content).decode('utf-8')
        
        return {
            'success': True,
            'filename': filename,
            'file_data': encoded_file
        }
        
    except Exception as e:
        logger.error(f"✗ Exception in api_generate_selected_devices_report: {type(e).__name__}: {str(e)}", exc_info=True)
        return {'success': False, 'message': str(e)}, 500


@app.route('/api/delete-devices-stream', methods=['POST'])
def api_delete_devices_stream():
    """Stream device deletion progress using SSE."""
    
    def generate():
        """Generator function for Server-Sent Events"""
        try:
            # Get device EUIs from request
            dev_euis = request.form.get('dev_euis', '')
            if not dev_euis:
                yield f"data: {json.dumps({'error': 'No devices specified'})}\n\n"
                return
            
            # Parse comma-separated dev_euis
            dev_eui_list = [eui.strip() for eui in dev_euis.split(',') if eui.strip()]
            total = len(dev_eui_list)
            
            logger.info(f"Bulk delete requested for {total} devices")
            
            # Send initial status
            yield f"data: {json.dumps({'status': 'starting', 'total': total, 'current': 0})}\n\n"
            
            # Create gRPC client
            from grpc_client import ChirpStackClient
            client = ChirpStackClient(SERVER_URL, API_CODE)
            connected, conn_msg = client.connect()
            
            if not connected:
                yield f"data: {json.dumps({'error': f'Connection failed: {conn_msg}'})}\n\n"
                return
            
            results = {'successful': [], 'failed': []}
            
            # Delete each device
            for idx, dev_eui in enumerate(dev_eui_list, 1):
                try:
                    logger.info(f"Deleting device {idx}/{total}: {dev_eui}")
                    
                    deleted, del_msg = client.delete_device(dev_eui)
                    
                    if deleted:
                        results['successful'].append({
                            'dev_eui': dev_eui
                        })
                        yield f"data: {json.dumps({'status': 'processing', 'current': idx, 'total': total, 'device': dev_eui, 'result': 'success'})}\n\n"
                    else:
                        results['failed'].append({
                            'dev_eui': dev_eui,
                            'error': del_msg
                        })
                        yield f"data: {json.dumps({'status': 'processing', 'current': idx, 'total': total, 'device': dev_eui, 'result': 'failed', 'message': del_msg})}\n\n"
                    
                    time.sleep(0.1)  # Small delay to avoid overwhelming the server
                    
                except Exception as e:
                    logger.error(f"Error deleting device {dev_eui}: {e}")
                    results['failed'].append({
                        'dev_eui': dev_eui,
                        'error': str(e)
                    })
                    yield f"data: {json.dumps({'status': 'processing', 'current': idx, 'total': total, 'device': dev_eui, 'result': 'failed', 'message': str(e)})}\n\n"
            
            # Close connection
            client.close()
            
            # Send completion
            logger.info(f"Bulk delete completed: {len(results['successful'])} successful, {len(results['failed'])} failed")
            yield f"data: {json.dumps({'status': 'complete', 'results': results})}\n\n"
            
        except Exception as e:
            logger.error(f"Error in delete stream: {e}", exc_info=True)
            yield f"data: {json.dumps({'error': str(e)})}\n\n"
    
    return Response(stream_with_context(generate()), content_type='text/event-stream')


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000, use_reloader=False)
